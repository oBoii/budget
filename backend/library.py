import json
import matplotlib.pyplot as plt
from root import ROOT
import os
from openpyxl import Workbook, load_workbook


def param(name):
    with open(ROOT + "params.json") as f:
        data = json.load(f)
        return data[name]


def append_expense(price, ratio, category, now):
    # append to csv
    with open(ROOT + "/expenses.csv", "a") as f:
        f.write(f"{now},{price},{ratio},{category}\n")


def append_expense_excel(price, ratio, category, name, now):
    # save columns: date, price, ratio, category, name

    ROOT = os.path.dirname(
        os.path.abspath(__file__)
    )  # Get the root directory of the script
    excel_file_path = os.path.join(ROOT, "expenses.xlsx")

    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Price", "Ratio", "Category", "Name"])
        wb.save(excel_file_path)

    wb = load_workbook(excel_file_path)
    ws = wb.active

    ws.append([now, price, ratio, category, name])
    wb.save(excel_file_path)


def get_debt_per_person():
    # Get the debt for Fabian and Elisa

    ROOT = os.path.dirname(
        os.path.abspath(__file__)
    )  # Get the root directory of the script
    excel_file_path = os.path.join(ROOT, "expenses.xlsx")

    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Price", "Ratio", "Category", "Name"])
        wb.save(excel_file_path)

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # get list of expenses
    expenses = []
    ratios = []
    payed_by_name = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        expense = row[1]
        ratio = row[2]
        payed_by = row[4]
        if expense is None:
            continue

        print(expense, ratio, payed_by)
        expenses.append(expense)
        ratios.append(ratio)
        payed_by_name.append(payed_by)

    # eg: Fabian paid 40 EUR with ratio 60% so he payed 40 * (1 - 0.6) = 16 EUR for Elisa
    amt_fabian_payed_for_elisa = sum(
        [
            expenses[i] * (1 - ratios[i] / 100)
            for i in range(len(expenses))
            if payed_by_name[i].lower() == "fabian"
        ]
    )

    amt_elisa_payed_for_fabian = sum(
        [
            expenses[i] * (1 - ratios[i] / 100)
            for i in range(len(expenses))
            if payed_by_name[i].lower() == "elisa"
        ]
    )

    fabian = amt_fabian_payed_for_elisa - amt_elisa_payed_for_fabian
    elisa = -fabian

    return fabian, elisa


# this code has a bug, eg: if you have 2 the same categories on same date but another category in between, it will not group them
# def group_expenses_by_category_and_date(data):
#     # data is a list of dicts, already sorted by date
#     # group by category and date and sum up the prices
#     # return a list of dicts

#     # list of dicts
#     data_grouped_by_category_and_date = []

#     previous_category = None
#     previous_date = None
#     for entry in data:
#         category = entry["category"]
#         date = entry["date"]

#         if category == previous_category and date == previous_date:
#             # add to previous entry
#             data_grouped_by_category_and_date[-1]["price"] += entry["price"]
#         else:
#             # create new entry
#             data_grouped_by_category_and_date.append(entry)

#         previous_category = category
#         previous_date = date

#     return data_grouped_by_category_and_date


# this function fixes the bug of the above function
def group_expenses_by_category_and_date(data):
    # list of dicts
    data_grouped_by_category_and_date = []

    occupied_categories = (
        []
    )  # the categories that are already used for a given date. If a category is already used for a given date, we need to add the price to the existing entry
    current_date = None
    for entry in data:
        category = entry["category"]
        date = entry["date"]

        if date != current_date:
            # new entry
            data_grouped_by_category_and_date.append(entry)
            current_date = date
            occupied_categories = [category]
        else:
            if category in occupied_categories:
                # add to existing entry

                # iterate over all entries in reverse order
                for i in range(len(data_grouped_by_category_and_date) - 1, -1, -1):
                    if (
                        data_grouped_by_category_and_date[i]["category"] == category
                        and data_grouped_by_category_and_date[i]["date"] == date
                    ):
                        # add price to this entry
                        data_grouped_by_category_and_date[i]["price"] += entry["price"]
                        break

            else:
                # create new entry
                data_grouped_by_category_and_date.append(entry)
                occupied_categories.append(category)

    return data_grouped_by_category_and_date


def get_expenses(name):
    # Get the expenses for `name`

    ROOT = os.path.dirname(
        os.path.abspath(__file__)
    )  # Get the root directory of the script
    excel_file_path = os.path.join(ROOT, "expenses.xlsx")

    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Price", "Ratio", "Category", "Name"])
        wb.save(excel_file_path)

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Get list of costs for `name`
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        date = row[0]
        expense = row[1]
        ratio = row[2]
        category = row[3]
        payed_by = row[4]

        if expense is None:
            continue

        individual_cost = {
            "date": date.strftime("%d-%m"),
            "price": 0,
            "category": category,
        }
        # if payed_by.lower() == name.lower():
        individual_cost["price"] = (
            expense * (ratio / 100)
            if payed_by.lower() == name.lower()
            else expense * (1 - ratio / 100)
        )

        # round to 2 decimals
        individual_cost["price"] = round(individual_cost["price"], 2)

        if individual_cost["price"] > 0:
            data.append(individual_cost)

    data = group_expenses_by_category_and_date(data)

    # reverse order
    data = data[::-1]

    # only first 30 entries
    data = data[:30]

    return data
